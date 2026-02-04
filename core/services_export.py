"""
Export Service Classes for Financial Management System

This module provides comprehensive export functionality for financial data
including PDF, Excel, and CSV formats with professional formatting,
large dataset optimization, and progress tracking.
"""

import csv
import json
import io
import threading
import time
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Any, Optional, Union, Iterator, Callable

from django.http import HttpResponse, FileResponse, JsonResponse, StreamingHttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from django.db.models import QuerySet, Sum, Count, Avg
from django.core.cache import cache
from django.core.paginator import Paginator

# Import for PDF generation (we'll use reportlab)
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Import for Excel generation (we'll use openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    User, FeeStructure, StudentFee, FeePayment, Scholarship, ScholarshipRecipient,
    StaffPayroll, PayrollStructure, FinancialTransaction, SchoolSettings
)


class ExportProgressTracker:
    """
    Track progress of export operations for large datasets
    """
    
    def __init__(self, export_id: str = None):
        self.export_id = export_id or str(uuid.uuid4())
        self.cache_key = f'export_progress_{self.export_id}'
        self.cache_timeout = 3600  # 1 hour
    
    def update_progress(self, current: int, total: int, status: str = 'processing', message: str = ''):
        """Update export progress in cache"""
        progress_data = {
            'export_id': self.export_id,
            'current': current,
            'total': total,
            'percentage': int((current / total * 100)) if total > 0 else 0,
            'status': status,  # 'processing', 'completed', 'failed'
            'message': message,
            'updated_at': timezone.now().isoformat(),
        }
        cache.set(self.cache_key, progress_data, self.cache_timeout)
        return progress_data
    
    def get_progress(self) -> Dict:
        """Get current export progress"""
        return cache.get(self.cache_key, {
            'export_id': self.export_id,
            'current': 0,
            'total': 0,
            'percentage': 0,
            'status': 'not_found',
            'message': 'Export not found or expired',
            'updated_at': timezone.now().isoformat(),
        })
    
    def mark_completed(self, message: str = 'Export completed successfully'):
        """Mark export as completed"""
        progress = self.get_progress()
        progress['status'] = 'completed'
        progress['message'] = message
        progress['percentage'] = 100
        progress['updated_at'] = timezone.now().isoformat()
        cache.set(self.cache_key, progress, self.cache_timeout)
    
    def mark_failed(self, error_message: str):
        """Mark export as failed"""
        progress = self.get_progress()
        progress['status'] = 'failed'
        progress['message'] = error_message
        progress['updated_at'] = timezone.now().isoformat()
        cache.set(self.cache_key, progress, self.cache_timeout)
    
    def clear_progress(self):
        """Clear progress data from cache"""
        cache.delete(self.cache_key)


class LargeDatasetExporter:
    """
    Handle efficient export of large datasets using chunking and streaming
    """
    
    CHUNK_SIZE = 1000  # Process 1000 records at a time
    
    @staticmethod
    def chunk_queryset(queryset: QuerySet, chunk_size: int = None) -> Iterator[List]:
        """
        Yield chunks of queryset for efficient processing
        
        Args:
            queryset: Django QuerySet to chunk
            chunk_size: Number of records per chunk
        
        Yields:
            Lists of model instances
        """
        chunk_size = chunk_size or LargeDatasetExporter.CHUNK_SIZE
        paginator = Paginator(queryset, chunk_size)
        
        for page_num in paginator.page_range:
            page = paginator.page(page_num)
            yield list(page.object_list)
    
    @staticmethod
    def stream_csv_export(data_iterator: Iterator, headers: List[str], 
                         metadata: Dict = None) -> Iterator[str]:
        """
        Stream CSV data for large datasets
        
        Args:
            data_iterator: Iterator yielding data chunks
            headers: Column headers
            metadata: Export metadata
        
        Yields:
            CSV rows as strings
        """
        # Create a pseudo-buffer for CSV writer
        class Echo:
            def write(self, value):
                return value
        
        writer = csv.writer(Echo())
        
        # Yield metadata rows
        if metadata:
            if metadata.get('title'):
                yield writer.writerow([metadata['title']])
            yield writer.writerow([f"Generated on: {metadata.get('generated_at', timezone.now()).strftime('%Y-%m-%d %H:%M:%S')}"])
            yield writer.writerow([f"Generated by: {metadata.get('generated_by', 'System')}"])
            yield writer.writerow([f"School: {metadata.get('school_name', 'School Management System')}"])
            yield writer.writerow([f"Export ID: {metadata.get('export_id', 'N/A')}"])
            yield writer.writerow([f"Total Records: {metadata.get('total_records', 'Unknown')}"])
            yield writer.writerow([])
        
        # Yield header row
        yield writer.writerow(headers)
        
        # Yield data rows
        for chunk in data_iterator:
            for row_data in chunk:
                csv_row = []
                for header in headers:
                    value = row_data.get(header, '')
                    if isinstance(value, Decimal):
                        csv_row.append(float(value))
                    elif value is None:
                        csv_row.append('')
                    else:
                        csv_row.append(str(value))
                yield writer.writerow(csv_row)


class ExportService:
    """
    Main export service class that handles multiple export formats
    for financial data with professional formatting and metadata.
    """
    
    def __init__(self):
        self.supported_formats = ['pdf', 'excel', 'csv']
        self.metadata = {
            'generated_at': timezone.now(),
            'generated_by': None,
            'school_name': self._get_school_name(),
            'export_id': str(uuid.uuid4()),
            'version': '1.0',
        }
        self.progress_tracker = None
    
    def _get_school_name(self) -> str:
        """Get school name from settings"""
        try:
            school_settings = SchoolSettings.objects.first()
            return school_settings.school_name if school_settings else "School Management System"
        except:
            return "School Management System"
    
    def set_user(self, user: User) -> None:
        """Set the user who is generating the export"""
        self.metadata['generated_by'] = user.get_full_name() if user else None
        self.metadata['user_id'] = user.id if user else None
    
    def enable_progress_tracking(self, export_id: str = None) -> str:
        """
        Enable progress tracking for this export
        
        Args:
            export_id: Optional custom export ID
        
        Returns:
            Export ID for tracking
        """
        if export_id:
            self.metadata['export_id'] = export_id
        self.progress_tracker = ExportProgressTracker(self.metadata['export_id'])
        return self.metadata['export_id']
    
    def export_data(self, data: Union[QuerySet, List[Dict]], 
                   export_format: str, 
                   filename: str,
                   headers: List[str] = None,
                   title: str = None,
                   use_streaming: bool = False,
                   **kwargs) -> HttpResponse:
        """
        Main export method that delegates to specific format handlers
        
        Args:
            data: QuerySet or list of dictionaries to export
            export_format: 'pdf', 'excel', or 'csv'
            filename: Base filename without extension
            headers: Column headers for the export
            title: Title for the export document
            use_streaming: Use streaming for large datasets (CSV only)
            **kwargs: Additional parameters for specific exporters
        
        Returns:
            HttpResponse with the exported file
        """
        try:
            if export_format not in self.supported_formats:
                error_msg = f"Unsupported export format: {export_format}"
                if self.progress_tracker:
                    self.progress_tracker.mark_failed(error_msg)
                raise ValueError(error_msg)
            
            # Add title to metadata
            if title:
                self.metadata['title'] = title
            
            # Generate filename with timestamp
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            full_filename = f"{filename}_{timestamp}"
            
            # Determine if we should use streaming based on data size
            if isinstance(data, list):
                self.metadata['total_records'] = len(data)
            elif hasattr(data, 'count') and callable(getattr(data, 'count', None)):
                total_records = data.count()
                self.metadata['total_records'] = total_records
                
                # Auto-enable streaming for large datasets (>5000 records)
                if total_records > 5000 and export_format == 'csv':
                    use_streaming = True
            
            if use_streaming and export_format == 'csv':
                return self._export_csv_streaming(data, full_filename, headers, title, **kwargs)
            else:
                # Convert QuerySet to list of dictionaries if needed
                if hasattr(data, 'values'):
                    data = list(data.values())
                
                if export_format == 'csv':
                    return self._export_csv(data, full_filename, headers, title, **kwargs)
                elif export_format == 'excel':
                    return self._export_excel(data, full_filename, headers, title, **kwargs)
                elif export_format == 'pdf':
                    return self._export_pdf(data, full_filename, headers, title, **kwargs)
        except Exception as e:
            if self.progress_tracker:
                self.progress_tracker.mark_failed(str(e))
            raise
    
    def _export_csv(self, data: List[Dict], filename: str, 
                   headers: List[str] = None, title: str = None, **kwargs) -> HttpResponse:
        """Export data to CSV format with enhanced functionality and metadata"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Add comprehensive metadata header
        if title:
            writer.writerow([title])
        writer.writerow([f"Generated on: {self.metadata['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([f"Generated by: {self.metadata['generated_by'] or 'System'}"])
        writer.writerow([f"School: {self.metadata['school_name']}"])
        writer.writerow([f"Export ID: {self.metadata['export_id']}"])
        writer.writerow([f"Export Version: {self.metadata['version']}"])
        writer.writerow([f"Total Records: {len(data)}"])
        writer.writerow([f"Format: CSV"])
        writer.writerow([])  # Empty row for spacing
        
        if not data:
            writer.writerow(['No data available'])
            return response
        
        # Use provided headers or extract from first data row
        if not headers and data:
            headers = list(data[0].keys())
        
        if headers:
            writer.writerow(headers)
            
            # Write data rows with progress tracking
            total = len(data)
            for idx, row in enumerate(data):
                csv_row = []
                for header in headers:
                    value = row.get(header, '')
                    # Handle Decimal and other special types
                    if isinstance(value, Decimal):
                        csv_row.append(float(value))
                    elif value is None:
                        csv_row.append('')
                    else:
                        csv_row.append(str(value))
                writer.writerow(csv_row)
                
                # Update progress every 100 rows
                if self.progress_tracker and idx % 100 == 0:
                    self.progress_tracker.update_progress(
                        idx + 1, total, 'processing', 
                        f'Exporting row {idx + 1} of {total}'
                    )
            
            # Mark as completed
            if self.progress_tracker:
                self.progress_tracker.mark_completed(f'Successfully exported {total} records')
        
        return response
    
    def _export_csv_streaming(self, queryset: QuerySet, filename: str, 
                             headers: List[str] = None, title: str = None, **kwargs) -> StreamingHttpResponse:
        """
        Export large datasets using streaming to handle memory efficiently
        
        Args:
            queryset: Django QuerySet to export
            filename: Base filename without extension
            headers: Column headers
            title: Export title
        
        Returns:
            StreamingHttpResponse with CSV data
        """
        total_records = queryset.count()
        self.metadata['total_records'] = total_records
        
        # Initialize progress tracking
        if self.progress_tracker:
            self.progress_tracker.update_progress(0, total_records, 'processing', 'Starting export...')
        
        # Prepare data iterator with progress tracking
        def data_iterator_with_progress():
            processed = 0
            for chunk in LargeDatasetExporter.chunk_queryset(queryset):
                # Convert chunk to dictionaries
                chunk_data = []
                for obj in chunk:
                    if hasattr(obj, '__dict__'):
                        # Convert model instance to dict
                        row_dict = {}
                        for header in headers:
                            # Try to get attribute value
                            value = getattr(obj, header, None)
                            if callable(value):
                                value = value()
                            row_dict[header] = value
                        chunk_data.append(row_dict)
                    else:
                        chunk_data.append(obj)
                
                processed += len(chunk)
                
                # Update progress
                if self.progress_tracker:
                    self.progress_tracker.update_progress(
                        processed, total_records, 'processing',
                        f'Processed {processed} of {total_records} records'
                    )
                
                yield chunk_data
        
        # Create streaming response
        response = StreamingHttpResponse(
            LargeDatasetExporter.stream_csv_export(
                data_iterator_with_progress(),
                headers,
                self.metadata
            ),
            content_type='text/csv'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        # Mark as completed (will be called after streaming finishes)
        if self.progress_tracker:
            # Note: This happens before streaming completes, but it's the best we can do
            self.progress_tracker.update_progress(
                total_records, total_records, 'processing',
                'Streaming data to client...'
            )
        
        return response
    
    def _export_excel(self, data: List[Dict], filename: str, 
                     headers: List[str] = None, title: str = None, **kwargs) -> HttpResponse:
        """Export data to Excel format with professional formatting and enhanced metadata"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = title or "Export Data"
        
        # Set up styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        metadata_font = Font(italic=True, size=10)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Add title and comprehensive metadata
        if title:
            worksheet.cell(row=current_row, column=1, value=title)
            worksheet.cell(row=current_row, column=1).font = title_font
            current_row += 2
        
        worksheet.cell(row=current_row, column=1, value=f"Generated on: {self.metadata['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 1
        
        worksheet.cell(row=current_row, column=1, value=f"Generated by: {self.metadata['generated_by'] or 'System'}")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 1
        
        worksheet.cell(row=current_row, column=1, value=f"School: {self.metadata['school_name']}")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 1
        
        worksheet.cell(row=current_row, column=1, value=f"Export ID: {self.metadata['export_id']}")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 1
        
        worksheet.cell(row=current_row, column=1, value=f"Export Version: {self.metadata['version']}")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 1
        
        worksheet.cell(row=current_row, column=1, value=f"Total Records: {len(data)}")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 1
        
        worksheet.cell(row=current_row, column=1, value=f"Format: Excel (XLSX)")
        worksheet.cell(row=current_row, column=1).font = metadata_font
        current_row += 2
        
        if not data:
            worksheet.cell(row=current_row, column=1, value="No data available")
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            workbook.save(response)
            return response
        
        # Use provided headers or extract from first data row
        if not headers and data:
            headers = list(data[0].keys())
        
        if headers:
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            current_row += 1
            
            # Write data rows with progress tracking
            total = len(data)
            for idx, row_data in enumerate(data):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    
                    # Handle different data types
                    if isinstance(value, Decimal):
                        value = float(value)
                    elif value is None:
                        value = ''
                    
                    cell = worksheet.cell(row=current_row, column=col_num, value=value)
                    cell.border = border
                    
                    # Format numbers
                    if isinstance(value, (int, float)) and header.lower() in ['amount', 'salary', 'fee', 'payment', 'balance']:
                        cell.number_format = '$#,##0.00'
                
                current_row += 1
                
                # Update progress every 100 rows
                if self.progress_tracker and idx % 100 == 0:
                    self.progress_tracker.update_progress(
                        idx + 1, total, 'processing',
                        f'Exporting row {idx + 1} of {total}'
                    )
            
            # Mark as completed
            if self.progress_tracker:
                self.progress_tracker.mark_completed(f'Successfully exported {total} records')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        workbook.save(response)
        return response
    
    def _export_pdf(self, data: List[Dict], filename: str, 
                   headers: List[str] = None, title: str = None, **kwargs) -> HttpResponse:
        """Export data to PDF format with professional formatting and enhanced metadata"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.darkblue
        )
        
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=6
        )
        
        # Add title
        if title:
            story.append(Paragraph(title, title_style))
        
        # Add comprehensive metadata
        story.append(Paragraph(f"<b>School:</b> {self.metadata['school_name']}", metadata_style))
        story.append(Paragraph(f"<b>Generated on:</b> {self.metadata['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}", metadata_style))
        story.append(Paragraph(f"<b>Generated by:</b> {self.metadata['generated_by'] or 'System'}", metadata_style))
        story.append(Paragraph(f"<b>Export ID:</b> {self.metadata['export_id']}", metadata_style))
        story.append(Paragraph(f"<b>Export Version:</b> {self.metadata['version']}", metadata_style))
        story.append(Paragraph(f"<b>Total Records:</b> {len(data)}", metadata_style))
        story.append(Paragraph(f"<b>Format:</b> PDF", metadata_style))
        story.append(Spacer(1, 20))
        
        if not data:
            story.append(Paragraph("No data available", styles['Normal']))
            doc.build(story)
            return response
        
        # Use provided headers or extract from first data row
        if not headers and data:
            headers = list(data[0].keys())
        
        if headers and data:
            # Prepare table data with progress tracking
            table_data = [headers]  # Header row
            
            total = len(data)
            for idx, row_data in enumerate(data):
                row = []
                for header in headers:
                    value = row_data.get(header, '')
                    
                    # Format values
                    if isinstance(value, Decimal):
                        if header.lower() in ['amount', 'salary', 'fee', 'payment', 'balance']:
                            row.append(f"${float(value):,.2f}")
                        else:
                            row.append(str(float(value)))
                    elif value is None:
                        row.append('')
                    else:
                        row.append(str(value))
                
                table_data.append(row)
                
                # Update progress every 100 rows
                if self.progress_tracker and idx % 100 == 0:
                    self.progress_tracker.update_progress(
                        idx + 1, total, 'processing',
                        f'Exporting row {idx + 1} of {total}'
                    )
            
            # Mark as completed
            if self.progress_tracker:
                self.progress_tracker.mark_completed(f'Successfully exported {total} records')
            
            # Create table
            table = Table(table_data)
            
            # Table styling
            table.setStyle(TableStyle([
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                
                # Data rows styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                
                # Grid and borders
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
        
        # Build PDF
        doc.build(story)
        return response


class ReceiptGenerator:
    """
    Specialized service for generating printable receipts for fee payments
    """
    
    def __init__(self):
        self.school_settings = self._get_school_settings()
    
    def _get_school_settings(self):
        """Get school settings for branding"""
        try:
            return SchoolSettings.objects.first()
        except:
            return None
    
    def generate_payment_receipt(self, payment: FeePayment, format: str = 'pdf') -> HttpResponse:
        """
        Generate a professional payment receipt
        
        Args:
            payment: FeePayment instance
            format: 'pdf' or 'html'
        
        Returns:
            HttpResponse with the receipt
        """
        if format == 'pdf':
            return self._generate_pdf_receipt(payment)
        elif format == 'html':
            return self._generate_html_receipt(payment)
        else:
            raise ValueError(f"Unsupported receipt format: {format}")
    
    def _generate_pdf_receipt(self, payment: FeePayment) -> HttpResponse:
        """Generate PDF receipt for payment"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF receipts. Install with: pip install reportlab")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{payment.id}_{payment.payment_date.strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        header_style = ParagraphStyle(
            'ReceiptHeader',
            parent=styles['Heading1'],
            fontSize=20,
            alignment=1,
            textColor=colors.darkblue,
            spaceAfter=10
        )
        
        school_style = ParagraphStyle(
            'SchoolInfo',
            parent=styles['Normal'],
            fontSize=12,
            alignment=1,
            spaceAfter=20
        )
        
        receipt_title_style = ParagraphStyle(
            'ReceiptTitle',
            parent=styles['Heading2'],
            fontSize=16,
            alignment=1,
            textColor=colors.darkgreen,
            spaceAfter=20
        )
        
        # School header
        if self.school_settings:
            story.append(Paragraph(self.school_settings.school_name, header_style))
            if self.school_settings.school_address:
                story.append(Paragraph(self.school_settings.school_address, school_style))
        else:
            story.append(Paragraph("School Management System", header_style))
        
        story.append(Paragraph("PAYMENT RECEIPT", receipt_title_style))
        
        # Receipt details
        receipt_data = [
            ['Receipt No:', f"RCP-{payment.id:06d}"],
            ['Date:', payment.payment_date.strftime('%B %d, %Y')],
            ['Time:', payment.payment_date.strftime('%I:%M %p')],
            ['', ''],
            ['Student Name:', payment.student_fee.student.user.get_full_name()],
            ['Student ID:', payment.student_fee.student.student_id],
            ['Class:', str(payment.student_fee.student.school_class)],
            ['', ''],
            ['Fee Type:', payment.student_fee.fee_structure.name],
            ['Term:', str(payment.student_fee.fee_structure.term)],
            ['Payment Method:', payment.get_payment_method_display()],
            ['Reference No:', payment.reference_number or 'N/A'],
            ['', ''],
            ['Amount Paid:', f"${payment.amount:,.2f}"],
            ['Received By:', payment.received_by.get_full_name() if payment.received_by else 'System'],
        ]
        
        # Add balance information
        balance = payment.student_fee.balance_amount
        if balance > 0:
            receipt_data.append(['Outstanding Balance:', f"${balance:,.2f}"])
        else:
            receipt_data.append(['Status:', 'FULLY PAID'])
        
        receipt_table = Table(receipt_data, colWidths=[2*inch, 3*inch])
        receipt_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            # Highlight amount paid
            ('BACKGROUND', (0, -2), (-1, -2), colors.lightgreen),
            ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
        ]))
        
        story.append(receipt_table)
        story.append(Spacer(1, 30))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.grey
        )
        
        story.append(Paragraph("Thank you for your payment!", footer_style))
        story.append(Paragraph("This is a computer-generated receipt.", footer_style))
        
        doc.build(story)
        return response
    
    def _generate_html_receipt(self, payment: FeePayment) -> HttpResponse:
        """Generate HTML receipt for payment (for printing)"""
        context = {
            'payment': payment,
            'school_settings': self.school_settings,
            'receipt_number': f"RCP-{payment.id:06d}",
            'balance': payment.student_fee.balance_amount,
        }
        
        html_content = render_to_string('financial/receipts/payment_receipt.html', context)
        return HttpResponse(html_content, content_type='text/html')


class PayrollSlipGenerator:
    """
    Specialized service for generating payroll slips for staff members
    """
    
    def __init__(self):
        self.school_settings = self._get_school_settings()
    
    def _get_school_settings(self):
        """Get school settings for branding"""
        try:
            return SchoolSettings.objects.first()
        except:
            return None
    
    def generate_payroll_slip(self, payroll: StaffPayroll, format: str = 'pdf') -> HttpResponse:
        """
        Generate a professional payroll slip
        
        Args:
            payroll: StaffPayroll instance
            format: 'pdf' or 'html'
        
        Returns:
            HttpResponse with the payroll slip
        """
        if format == 'pdf':
            return self._generate_pdf_payroll_slip(payroll)
        elif format == 'html':
            return self._generate_html_payroll_slip(payroll)
        else:
            raise ValueError(f"Unsupported payroll slip format: {format}")
    
    def _generate_pdf_payroll_slip(self, payroll: StaffPayroll) -> HttpResponse:
        """Generate PDF payroll slip"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF payroll slips. Install with: pip install reportlab")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payroll_slip_{payroll.teacher.employee_id}_{payroll.month.strftime("%Y_%m")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        header_style = ParagraphStyle(
            'PayrollHeader',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            textColor=colors.darkblue,
            spaceAfter=10
        )
        
        school_style = ParagraphStyle(
            'SchoolInfo',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            spaceAfter=20
        )
        
        slip_title_style = ParagraphStyle(
            'SlipTitle',
            parent=styles['Heading2'],
            fontSize=16,
            alignment=1,
            textColor=colors.darkgreen,
            spaceAfter=20
        )
        
        # School header
        if self.school_settings:
            story.append(Paragraph(self.school_settings.school_name, header_style))
            if self.school_settings.school_address:
                story.append(Paragraph(self.school_settings.school_address, school_style))
        else:
            story.append(Paragraph("School Management System", header_style))
        
        story.append(Paragraph("PAYROLL SLIP", slip_title_style))
        
        # Employee information
        employee_data = [
            ['Employee Name:', payroll.teacher.user.get_full_name()],
            ['Employee ID:', payroll.teacher.employee_id],
            ['Pay Period:', payroll.month.strftime('%B %Y')],
            ['Payroll Structure:', payroll.payroll_structure.name],
            ['Payment Date:', payroll.payment_date.strftime('%B %d, %Y') if payroll.payment_date else 'Pending'],
        ]
        
        employee_table = Table(employee_data, colWidths=[2*inch, 3*inch])
        employee_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(employee_table)
        story.append(Spacer(1, 20))
        
        # Earnings and deductions
        earnings_data = [
            ['EARNINGS', 'AMOUNT'],
            ['Basic Salary', f"${payroll.payroll_structure.basic_salary:,.2f}"],
            ['House Allowance', f"${payroll.payroll_structure.house_allowance:,.2f}"],
            ['Transport Allowance', f"${payroll.payroll_structure.transport_allowance:,.2f}"],
            ['Medical Allowance', f"${payroll.payroll_structure.medical_allowance:,.2f}"],
            ['Other Allowances', f"${payroll.payroll_structure.other_allowances:,.2f}"],
            ['', ''],
            ['GROSS SALARY', f"${payroll.gross_salary:,.2f}"],
        ]
        
        deductions_data = [
            ['DEDUCTIONS', 'AMOUNT'],
            ['Tax Deduction', f"${payroll.tax_deduction:,.2f}"],
            ['Pension Deduction', f"${payroll.pension_deduction:,.2f}"],
            ['Other Deductions', f"${payroll.other_deductions:,.2f}"],
            ['', ''],
            ['TOTAL DEDUCTIONS', f"${(payroll.tax_deduction + payroll.pension_deduction + payroll.other_deductions):,.2f}"],
            ['', ''],
            ['NET SALARY', f"${payroll.net_salary:,.2f}"],
        ]
        
        # Create side-by-side tables
        earnings_table = Table(earnings_data, colWidths=[1.8*inch, 1.2*inch])
        earnings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
        ]))
        
        deductions_table = Table(deductions_data, colWidths=[1.8*inch, 1.2*inch])
        deductions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightcoral),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
        ]))
        
        # Combine tables horizontally
        combined_table = Table([[earnings_table, deductions_table]], colWidths=[3*inch, 3*inch])
        combined_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(combined_table)
        story.append(Spacer(1, 30))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.grey
        )
        
        story.append(Paragraph("This is a computer-generated payroll slip.", footer_style))
        story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        doc.build(story)
        return response
    
    def _generate_html_payroll_slip(self, payroll: StaffPayroll) -> HttpResponse:
        """Generate HTML payroll slip (for printing)"""
        context = {
            'payroll': payroll,
            'school_settings': self.school_settings,
            'total_deductions': payroll.tax_deduction + payroll.pension_deduction + payroll.other_deductions,
        }
        
        html_content = render_to_string('financial/payroll/payroll_slip.html', context)
        return HttpResponse(html_content, content_type='text/html')
    
    def generate_bulk_payroll_slips(self, payrolls: QuerySet, format: str = 'pdf') -> HttpResponse:
        """
        Generate multiple payroll slips in a single file
        
        Args:
            payrolls: QuerySet of StaffPayroll instances
            format: 'pdf' only (HTML not supported for bulk)
        
        Returns:
            HttpResponse with combined payroll slips
        """
        if format != 'pdf':
            raise ValueError("Only PDF format is supported for bulk payroll slips")
        
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF payroll slips. Install with: pip install reportlab")
        
        response = HttpResponse(content_type='application/pdf')
        month_str = payrolls.first().month.strftime("%Y_%m") if payrolls.exists() else "unknown"
        response['Content-Disposition'] = f'attachment; filename="bulk_payroll_slips_{month_str}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        for i, payroll in enumerate(payrolls):
            if i > 0:
                # Add page break between payroll slips
                story.append(PageBreak())
            
            # Generate individual slip content (reuse logic from _generate_pdf_payroll_slip)
            # This is a simplified version - in practice, you'd extract the content generation
            # into a separate method to avoid duplication
            slip_content = self._get_payroll_slip_content(payroll)
            story.extend(slip_content)
        
        doc.build(story)
        return response
    
    def _get_payroll_slip_content(self, payroll: StaffPayroll) -> List:
        """Get the content elements for a single payroll slip"""
        # This method would contain the same logic as _generate_pdf_payroll_slip
        # but return the story elements instead of building the document
        # Implementation would be similar to the PDF generation above
        # but return a list of elements instead of building the document
        pass


class FinancialExportService:
    """
    Specialized export service for financial data with predefined formats
    and optimized handling for large datasets
    """
    
    def __init__(self):
        self.export_service = ExportService()
        self.receipt_generator = ReceiptGenerator()
        self.payroll_generator = PayrollSlipGenerator()
    
    def set_user(self, user: User) -> None:
        """Set the user for all export services"""
        self.export_service.set_user(user)
    
    def export_student_fees(self, queryset: QuerySet = None, 
                           export_format: str = 'csv', 
                           filters: Dict = None) -> HttpResponse:
        """Export student fees with optimized formatting"""
        if queryset is None:
            queryset = StudentFee.objects.all()
        
        if filters:
            # Apply filters if provided
            if filters.get('class_id'):
                queryset = queryset.filter(student__school_class_id=filters['class_id'])
            if filters.get('term_id'):
                queryset = queryset.filter(fee_structure__term_id=filters['term_id'])
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('date_from'):
                queryset = queryset.filter(created_at__gte=filters['date_from'])
            if filters.get('date_to'):
                queryset = queryset.filter(created_at__lte=filters['date_to'])
        
        # Prepare data with optimized queries
        data = []
        for fee in queryset.select_related('student__user', 'student__school_class', 'fee_structure__term'):
            data.append({
                'Student Name': fee.student.user.get_full_name(),
                'Student ID': fee.student.student_id,
                'Class': str(fee.student.school_class),
                'Fee Type': fee.fee_structure.name,
                'Term': str(fee.fee_structure.term),
                'Total Amount': fee.total_amount,
                'Paid Amount': fee.paid_amount,
                'Balance': fee.balance_amount,
                'Status': fee.get_status_display(),
                'Due Date': fee.due_date.strftime('%Y-%m-%d'),
                'Created Date': fee.created_at.strftime('%Y-%m-%d'),
            })
        
        headers = ['Student Name', 'Student ID', 'Class', 'Fee Type', 'Term', 
                  'Total Amount', 'Paid Amount', 'Balance', 'Status', 'Due Date', 'Created Date']
        
        return self.export_service.export_data(
            data=data,
            export_format=export_format,
            filename='student_fees_report',
            headers=headers,
            title='Student Fees Report'
        )
    
    def export_fee_payments(self, queryset: QuerySet = None, 
                           export_format: str = 'csv', 
                           filters: Dict = None) -> HttpResponse:
        """Export fee payments with detailed information"""
        if queryset is None:
            queryset = FeePayment.objects.all()
        
        if filters:
            if filters.get('date_from'):
                queryset = queryset.filter(payment_date__gte=filters['date_from'])
            if filters.get('date_to'):
                queryset = queryset.filter(payment_date__lte=filters['date_to'])
            if filters.get('payment_method'):
                queryset = queryset.filter(payment_method=filters['payment_method'])
            if filters.get('class_id'):
                queryset = queryset.filter(student_fee__student__school_class_id=filters['class_id'])
        
        data = []
        for payment in queryset.select_related(
            'student_fee__student__user', 
            'student_fee__student__school_class',
            'student_fee__fee_structure__term',
            'received_by'
        ):
            data.append({
                'Payment ID': payment.id,
                'Student Name': payment.student_fee.student.user.get_full_name(),
                'Student ID': payment.student_fee.student.student_id,
                'Class': str(payment.student_fee.student.school_class),
                'Fee Type': payment.student_fee.fee_structure.name,
                'Term': str(payment.student_fee.fee_structure.term),
                'Amount': payment.amount,
                'Payment Method': payment.get_payment_method_display(),
                'Reference Number': payment.reference_number or '',
                'Payment Date': payment.payment_date.strftime('%Y-%m-%d %H:%M:%S'),
                'Received By': payment.received_by.get_full_name() if payment.received_by else '',
                'Notes': payment.notes or '',
            })
        
        headers = ['Payment ID', 'Student Name', 'Student ID', 'Class', 'Fee Type', 'Term',
                  'Amount', 'Payment Method', 'Reference Number', 'Payment Date', 'Received By', 'Notes']
        
        return self.export_service.export_data(
            data=data,
            export_format=export_format,
            filename='fee_payments_report',
            headers=headers,
            title='Fee Payments Report'
        )
    
    def export_scholarships(self, queryset: QuerySet = None, 
                           export_format: str = 'csv', 
                           filters: Dict = None) -> HttpResponse:
        """Export scholarships with recipient information"""
        if queryset is None:
            queryset = Scholarship.objects.all()
        
        if filters:
            if filters.get('scholarship_type'):
                queryset = queryset.filter(scholarship_type=filters['scholarship_type'])
            if filters.get('academic_year'):
                queryset = queryset.filter(academic_year=filters['academic_year'])
            if filters.get('is_active') is not None:
                queryset = queryset.filter(is_active=filters['is_active'])
        
        data = []
        for scholarship in queryset.prefetch_related('scholarshiprecipient_set__student__user'):
            recipients = scholarship.scholarshiprecipient_set.filter(status='active')
            current_recipients = recipients.count()
            total_awarded = sum(r.awarded_amount for r in recipients)
            
            data.append({
                'Scholarship Name': scholarship.name,
                'Type': scholarship.get_scholarship_type_display(),
                'Academic Year': scholarship.academic_year,
                'Amount': scholarship.amount,
                'Percentage': scholarship.percentage or '',
                'Max Recipients': scholarship.max_recipients,
                'Current Recipients': current_recipients,
                'Total Awarded': total_awarded,
                'Is Active': 'Yes' if scholarship.is_active else 'No',
                'Created Date': scholarship.created_at.strftime('%Y-%m-%d'),
                'Description': scholarship.description or '',
            })
        
        headers = ['Scholarship Name', 'Type', 'Academic Year', 'Amount', 'Percentage',
                  'Max Recipients', 'Current Recipients', 'Total Awarded', 'Is Active', 
                  'Created Date', 'Description']
        
        return self.export_service.export_data(
            data=data,
            export_format=export_format,
            filename='scholarships_report',
            headers=headers,
            title='Scholarships Report'
        )
    
    def export_payroll(self, queryset: QuerySet = None, 
                      export_format: str = 'csv', 
                      filters: Dict = None) -> HttpResponse:
        """Export payroll data with detailed breakdown"""
        if queryset is None:
            queryset = StaffPayroll.objects.all()
        
        if filters:
            if filters.get('month'):
                queryset = queryset.filter(month=filters['month'])
            if filters.get('is_paid') is not None:
                queryset = queryset.filter(is_paid=filters['is_paid'])
            if filters.get('teacher_id'):
                queryset = queryset.filter(teacher_id=filters['teacher_id'])
        
        data = []
        for payroll in queryset.select_related('teacher__user', 'payroll_structure'):
            total_deductions = payroll.tax_deduction + payroll.pension_deduction + payroll.other_deductions
            
            data.append({
                'Employee Name': payroll.teacher.user.get_full_name(),
                'Employee ID': payroll.teacher.employee_id,
                'Month': payroll.month.strftime('%B %Y'),
                'Payroll Structure': payroll.payroll_structure.name,
                'Basic Salary': payroll.payroll_structure.basic_salary,
                'House Allowance': payroll.payroll_structure.house_allowance,
                'Transport Allowance': payroll.payroll_structure.transport_allowance,
                'Medical Allowance': payroll.payroll_structure.medical_allowance,
                'Other Allowances': payroll.payroll_structure.other_allowances,
                'Gross Salary': payroll.gross_salary,
                'Tax Deduction': payroll.tax_deduction,
                'Pension Deduction': payroll.pension_deduction,
                'Other Deductions': payroll.other_deductions,
                'Total Deductions': total_deductions,
                'Net Salary': payroll.net_salary,
                'Is Paid': 'Yes' if payroll.is_paid else 'No',
                'Payment Date': payroll.payment_date.strftime('%Y-%m-%d') if payroll.payment_date else '',
                'Created Date': payroll.created_at.strftime('%Y-%m-%d'),
            })
        
        headers = ['Employee Name', 'Employee ID', 'Month', 'Payroll Structure',
                  'Basic Salary', 'House Allowance', 'Transport Allowance', 'Medical Allowance',
                  'Other Allowances', 'Gross Salary', 'Tax Deduction', 'Pension Deduction',
                  'Other Deductions', 'Total Deductions', 'Net Salary', 'Is Paid', 
                  'Payment Date', 'Created Date']
        
        return self.export_service.export_data(
            data=data,
            export_format=export_format,
            filename='payroll_report',
            headers=headers,
            title='Payroll Report'
        )
    
    def export_financial_transactions(self, queryset: QuerySet = None, 
                                    export_format: str = 'csv', 
                                    filters: Dict = None) -> HttpResponse:
        """Export financial transactions"""
        if queryset is None:
            queryset = FinancialTransaction.objects.all()
        
        if filters:
            if filters.get('transaction_type'):
                queryset = queryset.filter(transaction_type=filters['transaction_type'])
            if filters.get('category'):
                queryset = queryset.filter(category=filters['category'])
            if filters.get('date_from'):
                queryset = queryset.filter(transaction_date__gte=filters['date_from'])
            if filters.get('date_to'):
                queryset = queryset.filter(transaction_date__lte=filters['date_to'])
        
        data = []
        for transaction in queryset.select_related('created_by'):
            data.append({
                'Transaction ID': transaction.id,
                'Type': transaction.get_transaction_type_display(),
                'Category': transaction.get_category_display(),
                'Amount': transaction.amount,
                'Description': transaction.description,
                'Reference Number': transaction.reference_number or '',
                'Transaction Date': transaction.transaction_date.strftime('%Y-%m-%d'),
                'Created By': transaction.created_by.get_full_name() if transaction.created_by else '',
                'Created Date': transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        headers = ['Transaction ID', 'Type', 'Category', 'Amount', 'Description',
                  'Reference Number', 'Transaction Date', 'Created By', 'Created Date']
        
        return self.export_service.export_data(
            data=data,
            export_format=export_format,
            filename='financial_transactions_report',
            headers=headers,
            title='Financial Transactions Report'
        )
    
    def generate_payment_receipt(self, payment_id: int, format: str = 'pdf') -> HttpResponse:
        """Generate payment receipt"""
        try:
            payment = FeePayment.objects.select_related(
                'student_fee__student__user',
                'student_fee__student__school_class',
                'student_fee__fee_structure__term',
                'received_by'
            ).get(id=payment_id)
            
            return self.receipt_generator.generate_payment_receipt(payment, format)
        except FeePayment.DoesNotExist:
            raise ValueError(f"Payment with ID {payment_id} not found")
    
    def generate_payroll_slip(self, payroll_id: int, format: str = 'pdf') -> HttpResponse:
        """Generate payroll slip"""
        try:
            payroll = StaffPayroll.objects.select_related(
                'teacher__user',
                'payroll_structure'
            ).get(id=payroll_id)
            
            return self.payroll_generator.generate_payroll_slip(payroll, format)
        except StaffPayroll.DoesNotExist:
            raise ValueError(f"Payroll with ID {payroll_id} not found")
    
    def generate_bulk_payroll_slips(self, month: str, format: str = 'pdf') -> HttpResponse:
        """Generate bulk payroll slips for a specific month"""
        try:
            from datetime import datetime
            month_date = datetime.strptime(month, '%Y-%m').date()
            
            payrolls = StaffPayroll.objects.filter(month=month_date).select_related(
                'teacher__user',
                'payroll_structure'
            )
            
            if not payrolls.exists():
                raise ValueError(f"No payroll records found for {month}")
            
            return self.payroll_generator.generate_bulk_payroll_slips(payrolls, format)
        except ValueError as e:
            if "time data" in str(e):
                raise ValueError("Invalid month format. Use YYYY-MM format (e.g., 2024-01)")
            raise


# Utility functions for backward compatibility and easy access
def export_student_fees(export_format='csv', filters=None, user=None):
    """Utility function to export student fees"""
    service = FinancialExportService()
    if user:
        service.set_user(user)
    return service.export_student_fees(export_format=export_format, filters=filters)

def export_fee_payments(export_format='csv', filters=None, user=None):
    """Utility function to export fee payments"""
    service = FinancialExportService()
    if user:
        service.set_user(user)
    return service.export_fee_payments(export_format=export_format, filters=filters)

def generate_payment_receipt(payment_id, format='pdf'):
    """Utility function to generate payment receipt"""
    service = FinancialExportService()
    return service.generate_payment_receipt(payment_id, format)

def generate_payroll_slip(payroll_id, format='pdf'):
    """Utility function to generate payroll slip"""
    service = FinancialExportService()
    return service.generate_payroll_slip(payroll_id, format)